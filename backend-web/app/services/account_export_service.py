"""
账号数据导出服务

功能：
1. 按勾选的账号ID或筛选条件查询账号及关联数据
2. 生成多Sheet的Excel文件（openpyxl）
3. 所有字段以字符串形式写入，避免数字类型导致精度丢失或 .0 后缀
"""
from __future__ import annotations

import io
import json
from typing import Optional

from loguru import logger
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession

from common.models.xy_account import XYAccount
from common.models.card import Card
from common.models.card_item_relation import CardItemRelation
from common.models.xy_keyword_rule import XYKeywordRule
from common.models.default_reply import DefaultReply
from common.models.xy_catalog_item import XYCatalogItem
from common.models.confirm_receipt_message import ConfirmReceiptMessage
from common.models.auto_rate_config import AutoRateConfig


def _to_str(value) -> str:
    """将任意值转为字符串，避免Excel自动转数字。None转为空字符串。"""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "是" if value else "否"
    if isinstance(value, float):
        # 避免 .0 后缀：如果是整数值就去掉小数部分
        if value == int(value):
            return str(int(value))
        return str(value)
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def _write_sheet(wb: Workbook, sheet_name: str, headers: list[str], rows: list[list]) -> None:
    """向工作簿写入一个Sheet，所有单元格设为文本格式。"""
    ws = wb.create_sheet(title=sheet_name)
    # 写表头
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    # 写数据行
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=_to_str(value))
            # 设置为文本格式，防止Excel自动转换
            cell.number_format = '@'
    # 自动调整列宽（简单估算）
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(headers[col_idx - 1]))
        for row_idx in range(2, min(len(rows) + 2, 50)):  # 只取前50行估算
            cell_val = ws.cell(row=row_idx, column=col_idx).value or ""
            max_len = max(max_len, min(len(str(cell_val)), 60))
        ws.column_dimensions[col_letter].width = max_len + 2


class AccountExportService:
    """账号数据导出服务"""

    def __init__(self, session: AsyncSession):
        self.session = session

    async def export_accounts(
        self,
        owner_id: int | None,
        account_ids: list[str] | None = None,
        filters: dict | None = None,
    ) -> io.BytesIO:
        """导出账号数据为Excel

        Args:
            owner_id: 所有者ID（None=管理员导出全部）
            account_ids: 指定导出的账号ID列表（勾选导出）
            filters: 筛选条件（未勾选时使用）

        Returns:
            Excel文件的BytesIO对象
        """
        # 1. 查询账号列表
        accounts = await self._get_accounts(owner_id, account_ids, filters)
        if not accounts:
            # 空数据也生成一个带表头的Excel
            pass

        account_id_list = [acc.account_id for acc in accounts]
        account_pk_list = [acc.id for acc in accounts]
        user_id_list = list({acc.owner_id for acc in accounts})

        # 2. 查询关联数据
        cards = await self._get_cards(user_id_list) if user_id_list else []
        card_ids = [c.id for c in cards]
        # 构建 card_id -> card_name 映射（卡券商品关联Sheet用名称代替主键）
        self._card_id_to_name = {c.id: c.name for c in cards}
        card_item_relations = await self._get_card_item_relations(card_ids) if card_ids else []
        keyword_rules = await self._get_keyword_rules(account_pk_list) if account_pk_list else []
        default_replies = await self._get_default_replies(account_id_list) if account_id_list else []
        message_filters = await self._get_message_filters(account_id_list) if account_id_list else []
        catalog_items = await self._get_catalog_items(account_pk_list) if account_pk_list else []
        confirm_receipt_msgs = await self._get_confirm_receipt_messages(account_id_list) if account_id_list else []
        auto_rate_configs = await self._get_auto_rate_configs(account_id_list) if account_id_list else []

        # 3. 构建 account_pk -> account_id 映射（关键词规则等用PK关联）
        pk_to_account_id = {acc.id: acc.account_id for acc in accounts}

        # 4. 生成Excel
        wb = Workbook()
        # 删除默认Sheet
        wb.remove(wb.active)

        self._write_account_basic(wb, accounts)
        self._write_account_switches(wb, accounts)
        self._write_ai_settings(wb, accounts)
        self._write_catalog_items(wb, catalog_items, pk_to_account_id)
        self._write_cards(wb, cards)
        self._write_card_item_relations(wb, card_item_relations)
        self._write_keyword_rules(wb, keyword_rules, pk_to_account_id)
        self._write_default_replies(wb, default_replies)
        self._write_message_filters(wb, message_filters)
        self._write_confirm_receipt(wb, confirm_receipt_msgs)
        self._write_auto_rate(wb, auto_rate_configs)

        # 5. 输出到BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    # ==================== 数据查询 ====================

    async def _get_accounts(
        self,
        owner_id: int | None,
        account_ids: list[str] | None,
        filters: dict | None,
    ) -> list[XYAccount]:
        """查询账号列表"""
        stmt = select(XYAccount)
        if owner_id is not None:
            stmt = stmt.where(XYAccount.owner_id == owner_id)
        if account_ids:
            stmt = stmt.where(XYAccount.account_id.in_(account_ids))
        elif filters:
            if filters.get("status"):
                stmt = stmt.where(XYAccount.status == filters["status"])
            if filters.get("account_id"):
                stmt = stmt.where(XYAccount.account_id.like(f"%{filters['account_id']}%"))
            if filters.get("has_password") is True:
                stmt = stmt.where(XYAccount.login_password.is_not(None))
                stmt = stmt.where(XYAccount.login_password != "")
            elif filters.get("has_password") is False:
                stmt = stmt.where(
                    (XYAccount.login_password.is_(None)) | (XYAccount.login_password == "")
                )
        stmt = stmt.order_by(XYAccount.id.asc())
        result = await self.session.execute(stmt)
        return list(result.scalars().all())

    async def _get_cards(self, user_ids: list[int]) -> list[Card]:
        stmt = select(Card).where(Card.user_id.in_(user_ids)).order_by(Card.id.asc())
        result = await self.session.execute(stmt)
        return list(result.scalars().all())

    async def _get_card_item_relations(self, card_ids: list[int]) -> list[CardItemRelation]:
        stmt = select(CardItemRelation).where(CardItemRelation.card_id.in_(card_ids))
        result = await self.session.execute(stmt)
        return list(result.scalars().all())

    async def _get_keyword_rules(self, account_pks: list[int]) -> list[XYKeywordRule]:
        stmt = select(XYKeywordRule).where(XYKeywordRule.account_pk.in_(account_pks)).order_by(XYKeywordRule.id.asc())
        result = await self.session.execute(stmt)
        return list(result.scalars().all())

    async def _get_default_replies(self, account_ids: list[str]) -> list[DefaultReply]:
        stmt = select(DefaultReply).where(DefaultReply.account_id.in_(account_ids)).order_by(DefaultReply.id.asc())
        result = await self.session.execute(stmt)
        return list(result.scalars().all())

    async def _get_message_filters(self, account_ids: list[str]) -> list:
        # 使用动态占位符避免 SQLAlchemy text() 不支持 IN tuple 的问题
        placeholders = ", ".join([f":id_{i}" for i in range(len(account_ids))])
        params = {f"id_{i}": aid for i, aid in enumerate(account_ids)}
        sql = text(f"""
            SELECT id, account_id, keyword, filter_type, enabled
            FROM xy_message_filters
            WHERE account_id IN ({placeholders})
            ORDER BY id
        """)
        result = await self.session.execute(sql, params)
        return result.fetchall()

    async def _get_catalog_items(self, account_pks: list[int]) -> list[XYCatalogItem]:
        stmt = select(XYCatalogItem).where(XYCatalogItem.account_pk.in_(account_pks)).order_by(XYCatalogItem.id.asc())
        result = await self.session.execute(stmt)
        return list(result.scalars().all())

    async def _get_confirm_receipt_messages(self, account_ids: list[str]) -> list[ConfirmReceiptMessage]:
        stmt = select(ConfirmReceiptMessage).where(ConfirmReceiptMessage.account_id.in_(account_ids))
        result = await self.session.execute(stmt)
        return list(result.scalars().all())

    async def _get_auto_rate_configs(self, account_ids: list[str]) -> list[AutoRateConfig]:
        stmt = select(AutoRateConfig).where(AutoRateConfig.account_id.in_(account_ids))
        result = await self.session.execute(stmt)
        return list(result.scalars().all())

    # ==================== Sheet写入 ====================

    def _write_account_basic(self, wb: Workbook, accounts: list[XYAccount]) -> None:
        headers = [
            "账号ID", "备注", "用户名", "登录密码", "Cookie", "状态", "禁用原因",
            "暂停时长(秒)", "相同消息等待时间(秒)", "显示浏览器",
            "代理类型", "代理地址", "代理端口", "代理用户名", "代理密码",
        ]
        rows = []
        for acc in accounts:
            rows.append([
                acc.account_id, acc.remark, acc.username, acc.login_password,
                acc.cookie, acc.status, acc.disable_reason,
                acc.pause_duration, acc.message_expire_time, acc.show_browser,
                acc.proxy_type, acc.proxy_host, acc.proxy_port, acc.proxy_user, acc.proxy_pass,
            ])
        _write_sheet(wb, "账号基本信息", headers, rows)

    def _write_account_switches(self, wb: Workbook, accounts: list[XYAccount]) -> None:
        headers = [
            "账号ID", "自动确认收货", "定时补发货", "定时补评价", "商品擦亮",
            "发货成功再发卡券", "自动求小红花", "禁止发货", "禁止发货原因",
            "主动关闭订单", "关闭后发卡券", "禁止发货排除商品",
        ]
        rows = []
        for acc in accounts:
            rows.append([
                acc.account_id, acc.auto_confirm, acc.scheduled_redelivery,
                acc.scheduled_rate, acc.auto_polish, acc.confirm_before_send,
                acc.auto_red_flower, acc.delivery_disabled, acc.delivery_disabled_reason,
                acc.auto_close_order, acc.delivery_only_card_after_close,
                acc.delivery_disabled_excluded_items,
            ])
        _write_sheet(wb, "账号开关配置", headers, rows)

    def _write_ai_settings(self, wb: Workbook, accounts: list[XYAccount]) -> None:
        headers = ["账号ID", "AI回复设置JSON"]
        rows = []
        for acc in accounts:
            ai_settings = (acc.metadata_json or {}).get("ai_reply_settings")
            rows.append([acc.account_id, ai_settings])
        _write_sheet(wb, "AI回复设置", headers, rows)

    def _write_catalog_items(self, wb: Workbook, items: list[XYCatalogItem], pk_map: dict) -> None:
        headers = ["账号ID", "商品ID", "标题", "价格", "AI提示词"]
        rows = []
        for item in items:
            rows.append([
                pk_map.get(item.account_pk, ""),
                item.item_id, item.title, item.price, item.ai_prompt,
            ])
        _write_sheet(wb, "商品目录", headers, rows)

    def _write_cards(self, wb: Workbook, cards: list[Card]) -> None:
        headers = [
            "卡券名称", "类型", "启用", "延迟秒数",
            "多规格", "规格名", "规格值",
            "API配置", "文本内容", "数据内容", "图片URL", "多图片URL",
        ]
        rows = []
        for card in cards:
            rows.append([
                card.name, card.type, card.enabled, card.delay_seconds,
                card.is_multi_spec, card.spec_name, card.spec_value,
                card.api_config, card.text_content, card.data_content,
                card.image_url, card.image_urls,
            ])
        _write_sheet(wb, "卡券", headers, rows)

    def _write_card_item_relations(self, wb: Workbook, relations: list[CardItemRelation]) -> None:
        headers = ["卡券名称", "商品ID", "来源"]
        rows = []
        for rel in relations:
            card_name = self._card_id_to_name.get(rel.card_id, "")
            rows.append([card_name, rel.item_id, rel.source])
        _write_sheet(wb, "卡券商品关联", headers, rows)

    def _write_keyword_rules(self, wb: Workbook, rules: list[XYKeywordRule], pk_map: dict) -> None:
        headers = ["账号ID", "关键词", "回复内容", "回复类型", "图片URL", "商品ID", "优先级", "启用"]
        rows = []
        for rule in rules:
            rows.append([
                pk_map.get(rule.account_pk, ""),
                rule.keyword, rule.reply_content, rule.reply_type,
                rule.image_url, rule.item_id, rule.priority, rule.is_active,
            ])
        _write_sheet(wb, "关键词规则", headers, rows)

    def _write_default_replies(self, wb: Workbook, replies: list[DefaultReply]) -> None:
        headers = ["账号ID", "商品ID", "启用", "回复内容", "回复图片", "仅回复一次", "回复类型", "API地址", "API超时"]
        rows = []
        for reply in replies:
            rows.append([
                reply.account_id, reply.item_id, reply.enabled,
                reply.reply_content, reply.reply_image, reply.reply_once,
                getattr(reply, "reply_type", "text") or "text",
                getattr(reply, "api_url", "") or "",
                getattr(reply, "api_timeout", 80) or 80,
            ])
        _write_sheet(wb, "默认回复", headers, rows)

    def _write_message_filters(self, wb: Workbook, filters: list) -> None:
        headers = ["账号ID", "关键词", "过滤类型", "启用"]
        rows = []
        for f in filters:
            rows.append([f.account_id, f.keyword, f.filter_type, bool(f.enabled)])
        _write_sheet(wb, "消息过滤规则", headers, rows)

    def _write_confirm_receipt(self, wb: Workbook, msgs: list[ConfirmReceiptMessage]) -> None:
        headers = ["账号ID", "启用", "消息内容", "消息图片"]
        rows = []
        for msg in msgs:
            rows.append([msg.account_id, msg.enabled, msg.message_content, msg.message_image])
        _write_sheet(wb, "确认收货消息", headers, rows)

    def _write_auto_rate(self, wb: Workbook, configs: list[AutoRateConfig]) -> None:
        headers = ["账号ID", "启用", "评价类型", "评价内容", "API地址"]
        rows = []
        for cfg in configs:
            rows.append([cfg.account_id, cfg.enabled, cfg.rate_type, cfg.text_content, cfg.api_url])
        _write_sheet(wb, "自动评价配置", headers, rows)
