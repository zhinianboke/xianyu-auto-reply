"""
账号数据导入服务

功能：
1. 解析导出的Excel文件（openpyxl）
2. 按Sheet逐一导入数据到数据库（upsert逻辑）
3. 支持两种模式：保存（按Excel状态）/ 保存并全部启用
4. 返回导入统计结果
"""
from __future__ import annotations

import json
from io import BytesIO
from typing import Any

from loguru import logger
from openpyxl import load_workbook
from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession

from common.models.xy_account import XYAccount
from common.models.card import Card
from common.models.card_item_relation import CardItemRelation
from common.models.xy_keyword_rule import XYKeywordRule
from common.models.default_reply import DefaultReply
from common.models.xy_catalog_item import XYCatalogItem
from common.models.confirm_receipt_message import ConfirmReceiptMessage
from common.models.auto_rate_config import AutoRateConfig
from common.utils.time_utils import get_beijing_now_naive


def _parse_bool(value: str | None) -> bool:
    """解析布尔值：'是'/True/'true'/'1' → True，其余 → False"""
    if value is None:
        return False
    v = str(value).strip().lower()
    return v in ("是", "true", "1", "yes")


def _parse_int(value: str | None, default: int = 0) -> int:
    """解析整数，失败返回默认值"""
    if value is None or str(value).strip() == "":
        return default
    try:
        return int(float(str(value).strip()))
    except (ValueError, TypeError):
        return default


def _parse_str(value: Any) -> str:
    """解析字符串，None → 空字符串"""
    if value is None:
        return ""
    return str(value).strip()


def _parse_json(value: str | None) -> Any:
    """解析JSON字符串，失败返回None"""
    if not value or str(value).strip() == "":
        return None
    try:
        return json.loads(str(value))
    except (json.JSONDecodeError, TypeError):
        return None


def _read_sheet_rows(wb, sheet_name: str) -> list[dict[str, str]]:
    """读取Sheet为字典列表（表头作为key）"""
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    headers = [str(h or "").strip() for h in rows[0]]
    result = []
    for row in rows[1:]:
        row_dict = {}
        for i, header in enumerate(headers):
            if not header:
                continue
            val = row[i] if i < len(row) else None
            row_dict[header] = str(val).strip() if val is not None else ""
        # 跳过全空行
        if any(v for v in row_dict.values()):
            result.append(row_dict)
    return result


class AccountImportService:
    """账号数据导入服务"""

    def __init__(self, session: AsyncSession, owner_id: int):
        self.session = session
        self.owner_id = owner_id
        self.inserted = 0
        self.updated = 0
        self.started = 0
        self.failed = 0
        self.errors: list[str] = []
        # 导入过程中的映射缓存
        self._account_id_to_pk: dict[str, int] = {}
        self._card_name_spec_to_id: dict[str, int] = {}

    async def import_accounts(
        self,
        file_content: bytes,
        enable_all: bool = False,
    ) -> dict:
        """导入账号数据

        Args:
            file_content: Excel文件内容
            enable_all: 是否全部启用

        Returns:
            导入结果统计
        """
        try:
            wb = load_workbook(BytesIO(file_content), read_only=True, data_only=True)
        except Exception as e:
            return {
                "success": False,
                "message": f"Excel文件解析失败: {str(e)}",
                "data": None,
            }

        try:
            # 按顺序导入各Sheet
            await self._import_accounts_basic(wb, enable_all)
            await self._import_account_switches(wb)
            await self._import_ai_settings(wb)
            await self._import_catalog_items(wb)
            await self._import_cards(wb)
            await self._import_card_item_relations(wb)
            await self._import_keyword_rules(wb)
            await self._import_default_replies(wb)
            await self._import_message_filters(wb)
            await self._import_confirm_receipt(wb)
            await self._import_auto_rate(wb)

            # 启动需要启用的账号
            accounts_to_start = await self._get_accounts_to_start(enable_all)
            for account_id, cookie in accounts_to_start:
                try:
                    await self._start_account(account_id, cookie)
                    self.started += 1
                except Exception as e:
                    self.errors.append(f"账号 {account_id} 启动失败: {str(e)}")

        except Exception as e:
            logger.error(f"导入过程异常: {e}")
            self.errors.append(f"导入异常: {str(e)}")

        message = (
            f"导入完成：新增 {self.inserted} 个，更新 {self.updated} 个，"
            f"启动 {self.started} 个，失败 {self.failed} 个"
        )
        return {
            "success": True,
            "message": message,
            "data": {
                "inserted": self.inserted,
                "updated": self.updated,
                "started": self.started,
                "failed": self.failed,
                "errors": self.errors[:20],  # 最多返回20条错误
            },
        }

    # ==================== 各Sheet导入逻辑 ====================

    async def _import_accounts_basic(self, wb, enable_all: bool) -> None:
        """导入账号基本信息"""
        rows = _read_sheet_rows(wb, "账号基本信息")
        for row in rows:
            account_id = _parse_str(row.get("账号ID"))
            if not account_id:
                continue

            cookie = _parse_str(row.get("Cookie"))
            if not cookie:
                self.errors.append(f"账号 {account_id}: Cookie为空，跳过")
                self.failed += 1
                continue

            # 确定状态
            status = _parse_str(row.get("状态")) or "active"
            if enable_all:
                status = "active"

            # 查询是否已存在
            stmt = select(XYAccount).where(
                XYAccount.account_id == account_id,
                XYAccount.owner_id == self.owner_id,
            )
            result = await self.session.execute(stmt)
            existing = result.scalars().first()

            if existing:
                # 更新
                existing.cookie = cookie
                existing.status = status
                existing.remark = _parse_str(row.get("备注")) or existing.remark
                existing.username = _parse_str(row.get("用户名")) or existing.username
                existing.login_password = _parse_str(row.get("登录密码")) or existing.login_password
                existing.pause_duration = _parse_int(row.get("暂停时长(秒)"), existing.pause_duration)
                existing.message_expire_time = _parse_int(row.get("相同消息等待时间(秒)"), existing.message_expire_time)
                existing.show_browser = _parse_bool(row.get("显示浏览器"))
                existing.proxy_type = _parse_str(row.get("代理类型")) or existing.proxy_type
                existing.proxy_host = _parse_str(row.get("代理地址")) or existing.proxy_host
                existing.proxy_port = _parse_int(row.get("代理端口"), existing.proxy_port or 0) or None
                existing.proxy_user = _parse_str(row.get("代理用户名")) or existing.proxy_user
                existing.proxy_pass = _parse_str(row.get("代理密码")) or existing.proxy_pass
                if enable_all:
                    existing.disable_reason = None
                self.session.add(existing)
                self._account_id_to_pk[account_id] = existing.id
                self.updated += 1
            else:
                # 新增
                account = XYAccount(
                    owner_id=self.owner_id,
                    account_id=account_id,
                    cookie=cookie,
                    login_method="import",
                    status=status,
                    remark=_parse_str(row.get("备注")),
                    username=_parse_str(row.get("用户名")),
                    login_password=_parse_str(row.get("登录密码")),
                    pause_duration=_parse_int(row.get("暂停时长(秒)"), 10),
                    message_expire_time=_parse_int(row.get("相同消息等待时间(秒)"), 3600),
                    show_browser=_parse_bool(row.get("显示浏览器")),
                    proxy_type=_parse_str(row.get("代理类型")) or "none",
                    proxy_host=_parse_str(row.get("代理地址")) or None,
                    proxy_port=_parse_int(row.get("代理端口"), 0) or None,
                    proxy_user=_parse_str(row.get("代理用户名")) or None,
                    proxy_pass=_parse_str(row.get("代理密码")) or None,
                )
                self.session.add(account)
                self.inserted += 1

            await self.session.flush()

            # 刷新后获取PK
            if account_id not in self._account_id_to_pk:
                stmt2 = select(XYAccount.id).where(
                    XYAccount.account_id == account_id,
                    XYAccount.owner_id == self.owner_id,
                )
                r = await self.session.execute(stmt2)
                pk = r.scalar_one_or_none()
                if pk:
                    self._account_id_to_pk[account_id] = pk

        await self.session.commit()

    async def _import_account_switches(self, wb) -> None:
        """导入账号开关配置"""
        rows = _read_sheet_rows(wb, "账号开关配置")
        for row in rows:
            account_id = _parse_str(row.get("账号ID"))
            if not account_id:
                continue
            stmt = select(XYAccount).where(
                XYAccount.account_id == account_id,
                XYAccount.owner_id == self.owner_id,
            )
            result = await self.session.execute(stmt)
            account = result.scalars().first()
            if not account:
                continue

            account.auto_confirm = _parse_bool(row.get("自动确认收货"))
            account.scheduled_redelivery = _parse_bool(row.get("定时补发货"))
            account.scheduled_rate = _parse_bool(row.get("定时补评价"))
            account.auto_polish = _parse_bool(row.get("商品擦亮"))
            account.confirm_before_send = _parse_bool(row.get("发货成功再发卡券"))
            account.auto_red_flower = _parse_bool(row.get("自动求小红花"))
            account.delivery_disabled = _parse_bool(row.get("禁止发货"))
            account.delivery_disabled_reason = _parse_str(row.get("禁止发货原因")) or None
            account.auto_close_order = _parse_bool(row.get("主动关闭订单"))
            account.delivery_only_card_after_close = _parse_bool(row.get("关闭后发卡券"))
            excluded = _parse_json(row.get("禁止发货排除商品"))
            if isinstance(excluded, list):
                account.delivery_disabled_excluded_items = excluded
            self.session.add(account)

        await self.session.commit()

    async def _import_ai_settings(self, wb) -> None:
        """导入AI回复设置"""
        rows = _read_sheet_rows(wb, "AI回复设置")
        for row in rows:
            account_id = _parse_str(row.get("账号ID"))
            if not account_id:
                continue
            ai_json = _parse_json(row.get("AI回复设置JSON"))
            if not ai_json:
                continue
            stmt = select(XYAccount).where(
                XYAccount.account_id == account_id,
                XYAccount.owner_id == self.owner_id,
            )
            result = await self.session.execute(stmt)
            account = result.scalars().first()
            if not account:
                continue

            metadata = account.metadata_json or {}
            metadata["ai_reply_settings"] = ai_json
            account.metadata_json = metadata
            self.session.add(account)

        await self.session.commit()

    async def _import_catalog_items(self, wb) -> None:
        """导入商品目录"""
        rows = _read_sheet_rows(wb, "商品目录")
        for row in rows:
            account_id = _parse_str(row.get("账号ID"))
            item_id = _parse_str(row.get("商品ID"))
            if not account_id or not item_id:
                continue
            account_pk = self._account_id_to_pk.get(account_id)
            if not account_pk:
                continue

            stmt = select(XYCatalogItem).where(
                XYCatalogItem.account_pk == account_pk,
                XYCatalogItem.item_id == item_id,
            )
            result = await self.session.execute(stmt)
            existing = result.scalars().first()

            if existing:
                existing.title = _parse_str(row.get("标题")) or existing.title
                existing.price = _parse_str(row.get("价格")) or existing.price
                existing.ai_prompt = _parse_str(row.get("AI提示词")) or existing.ai_prompt
            else:
                item = XYCatalogItem(
                    owner_id=self.owner_id,
                    account_pk=account_pk,
                    item_id=item_id,
                    title=_parse_str(row.get("标题")),
                    price=_parse_str(row.get("价格")),
                    ai_prompt=_parse_str(row.get("AI提示词")) or None,
                    created_at=get_beijing_now_naive(),
                )
                self.session.add(item)

        await self.session.commit()

    async def _import_cards(self, wb) -> None:
        """导入卡券"""
        rows = _read_sheet_rows(wb, "卡券")
        for row in rows:
            name = _parse_str(row.get("卡券名称"))
            if not name:
                continue
            card_type = _parse_str(row.get("类型")) or "text"
            spec_value = _parse_str(row.get("规格值"))

            # 按 名称+规格值 查找
            stmt = select(Card).where(
                Card.user_id == self.owner_id,
                Card.name == name,
            )
            if spec_value:
                stmt = stmt.where(Card.spec_value == spec_value)
            else:
                stmt = stmt.where((Card.spec_value.is_(None)) | (Card.spec_value == ""))

            result = await self.session.execute(stmt)
            existing = result.scalars().first()

            if existing:
                existing.type = card_type
                existing.enabled = _parse_bool(row.get("启用"))
                existing.delay_seconds = _parse_int(row.get("延迟秒数"), 0)
                existing.is_multi_spec = _parse_bool(row.get("多规格"))
                existing.spec_name = _parse_str(row.get("规格名")) or existing.spec_name
                existing.spec_value = spec_value or existing.spec_value
                existing.api_config = _parse_str(row.get("API配置")) or existing.api_config
                existing.text_content = _parse_str(row.get("文本内容")) or existing.text_content
                existing.data_content = _parse_str(row.get("数据内容")) or existing.data_content
                existing.image_url = _parse_str(row.get("图片URL")) or existing.image_url
                existing.image_urls = _parse_str(row.get("多图片URL")) or existing.image_urls
                self.session.add(existing)
                self._card_name_spec_to_id[f"{name}|{spec_value}"] = existing.id
            else:
                card = Card(
                    user_id=self.owner_id,
                    name=name,
                    type=card_type,
                    enabled=_parse_bool(row.get("启用")),
                    delay_seconds=_parse_int(row.get("延迟秒数"), 0),
                    is_multi_spec=_parse_bool(row.get("多规格")),
                    spec_name=_parse_str(row.get("规格名")) or None,
                    spec_value=spec_value or None,
                    api_config=_parse_str(row.get("API配置")) or None,
                    text_content=_parse_str(row.get("文本内容")) or None,
                    data_content=_parse_str(row.get("数据内容")) or None,
                    image_url=_parse_str(row.get("图片URL")) or None,
                    image_urls=_parse_str(row.get("多图片URL")) or None,
                )
                self.session.add(card)
                await self.session.flush()
                self._card_name_spec_to_id[f"{name}|{spec_value}"] = card.id

        await self.session.commit()

        # 补充映射：查询所有该用户的卡券
        stmt = select(Card.id, Card.name, Card.spec_value).where(Card.user_id == self.owner_id)
        result = await self.session.execute(stmt)
        for card_id, card_name, card_spec in result.all():
            key = f"{card_name}|{card_spec or ''}"
            self._card_name_spec_to_id[key] = card_id

    async def _import_card_item_relations(self, wb) -> None:
        """导入卡券商品关联"""
        rows = _read_sheet_rows(wb, "卡券商品关联")
        for row in rows:
            card_name = _parse_str(row.get("卡券名称"))
            item_id = _parse_str(row.get("商品ID"))
            if not card_name or not item_id:
                continue

            # 查找卡券ID（先精确匹配名称，不带规格）
            card_id = None
            for key, cid in self._card_name_spec_to_id.items():
                if key.startswith(f"{card_name}|"):
                    card_id = cid
                    break
            if not card_id:
                continue

            # 检查是否已存在
            stmt = select(CardItemRelation).where(
                CardItemRelation.card_id == card_id,
                CardItemRelation.item_id == item_id,
                CardItemRelation.user_id == self.owner_id,
            )
            result = await self.session.execute(stmt)
            if result.scalars().first():
                continue  # 已存在，跳过

            source = _parse_str(row.get("来源")) or "own"
            rel = CardItemRelation(
                user_id=self.owner_id,
                card_id=card_id,
                item_id=item_id,
                source=source,
            )
            self.session.add(rel)

        await self.session.commit()

    async def _import_keyword_rules(self, wb) -> None:
        """导入关键词规则"""
        rows = _read_sheet_rows(wb, "关键词规则")
        for row in rows:
            account_id = _parse_str(row.get("账号ID"))
            keyword = _parse_str(row.get("关键词"))
            if not account_id or not keyword:
                continue
            account_pk = self._account_id_to_pk.get(account_id)
            if not account_pk:
                continue

            item_id = _parse_str(row.get("商品ID")) or None

            stmt = select(XYKeywordRule).where(
                XYKeywordRule.account_pk == account_pk,
                XYKeywordRule.keyword == keyword,
            )
            if item_id:
                stmt = stmt.where(XYKeywordRule.item_id == item_id)
            else:
                stmt = stmt.where((XYKeywordRule.item_id.is_(None)) | (XYKeywordRule.item_id == ""))

            result = await self.session.execute(stmt)
            existing = result.scalars().first()

            if existing:
                existing.reply_content = _parse_str(row.get("回复内容")) or existing.reply_content
                existing.reply_type = _parse_str(row.get("回复类型")) or existing.reply_type
                existing.image_url = _parse_str(row.get("图片URL")) or existing.image_url
                existing.priority = _parse_int(row.get("优先级"), existing.priority)
                existing.is_active = _parse_bool(row.get("启用"))
            else:
                rule = XYKeywordRule(
                    owner_id=self.owner_id,
                    account_pk=account_pk,
                    keyword=keyword,
                    reply_content=_parse_str(row.get("回复内容")) or None,
                    reply_type=_parse_str(row.get("回复类型")) or "text",
                    image_url=_parse_str(row.get("图片URL")) or None,
                    item_id=item_id,
                    priority=_parse_int(row.get("优先级"), 100),
                    is_active=_parse_bool(row.get("启用")),
                )
                self.session.add(rule)

        await self.session.commit()

    async def _import_default_replies(self, wb) -> None:
        """导入默认回复"""
        rows = _read_sheet_rows(wb, "默认回复")
        for row in rows:
            account_id = _parse_str(row.get("账号ID"))
            if not account_id:
                continue
            item_id = _parse_str(row.get("商品ID")) or None

            stmt = select(DefaultReply).where(DefaultReply.account_id == account_id)
            if item_id:
                stmt = stmt.where(DefaultReply.item_id == item_id)
            else:
                stmt = stmt.where((DefaultReply.item_id.is_(None)) | (DefaultReply.item_id == ""))

            result = await self.session.execute(stmt)
            existing = result.scalars().first()

            if existing:
                existing.enabled = _parse_bool(row.get("启用"))
                existing.reply_content = _parse_str(row.get("回复内容")) or existing.reply_content
                existing.reply_image = _parse_str(row.get("回复图片")) or existing.reply_image
                existing.reply_once = _parse_bool(row.get("仅回复一次"))
                existing.reply_type = _parse_str(row.get("回复类型")) or existing.reply_type or "text"
                existing.api_url = _parse_str(row.get("API地址")) or existing.api_url
                api_timeout = _parse_int(row.get("API超时"))
                if api_timeout:
                    existing.api_timeout = api_timeout
            else:
                reply = DefaultReply(
                    account_id=account_id,
                    item_id=item_id,
                    enabled=_parse_bool(row.get("启用")),
                    reply_content=_parse_str(row.get("回复内容")) or None,
                    reply_image=_parse_str(row.get("回复图片")) or None,
                    reply_once=_parse_bool(row.get("仅回复一次")),
                    reply_type=_parse_str(row.get("回复类型")) or "text",
                    api_url=_parse_str(row.get("API地址")) or None,
                    api_timeout=_parse_int(row.get("API超时")) or 80,
                )
                self.session.add(reply)

        await self.session.commit()

    async def _import_message_filters(self, wb) -> None:
        """导入消息过滤规则"""
        rows = _read_sheet_rows(wb, "消息过滤规则")
        for row in rows:
            account_id = _parse_str(row.get("账号ID"))
            keyword = _parse_str(row.get("关键词"))
            filter_type = _parse_str(row.get("过滤类型"))
            if not account_id or not keyword or not filter_type:
                continue

            # 检查唯一约束
            check_sql = text("""
                SELECT id FROM xy_message_filters
                WHERE account_id = :account_id AND keyword = :keyword AND filter_type = :filter_type
                LIMIT 1
            """)
            result = await self.session.execute(check_sql, {
                "account_id": account_id, "keyword": keyword, "filter_type": filter_type
            })
            if result.scalar_one_or_none():
                continue  # 已存在，跳过

            enabled = _parse_bool(row.get("启用"))
            insert_sql = text("""
                INSERT INTO xy_message_filters (account_id, keyword, filter_type, enabled)
                VALUES (:account_id, :keyword, :filter_type, :enabled)
            """)
            await self.session.execute(insert_sql, {
                "account_id": account_id, "keyword": keyword,
                "filter_type": filter_type, "enabled": 1 if enabled else 0,
            })

        await self.session.commit()

    async def _import_confirm_receipt(self, wb) -> None:
        """导入确认收货消息"""
        rows = _read_sheet_rows(wb, "确认收货消息")
        for row in rows:
            account_id = _parse_str(row.get("账号ID"))
            if not account_id:
                continue

            stmt = select(ConfirmReceiptMessage).where(ConfirmReceiptMessage.account_id == account_id)
            result = await self.session.execute(stmt)
            existing = result.scalars().first()

            if existing:
                existing.enabled = _parse_bool(row.get("启用"))
                existing.message_content = _parse_str(row.get("消息内容")) or existing.message_content
                existing.message_image = _parse_str(row.get("消息图片")) or existing.message_image
            else:
                msg = ConfirmReceiptMessage(
                    account_id=account_id,
                    enabled=_parse_bool(row.get("启用")),
                    message_content=_parse_str(row.get("消息内容")) or None,
                    message_image=_parse_str(row.get("消息图片")) or None,
                )
                self.session.add(msg)

        await self.session.commit()

    async def _import_auto_rate(self, wb) -> None:
        """导入自动评价配置"""
        rows = _read_sheet_rows(wb, "自动评价配置")
        for row in rows:
            account_id = _parse_str(row.get("账号ID"))
            if not account_id:
                continue

            stmt = select(AutoRateConfig).where(AutoRateConfig.account_id == account_id)
            result = await self.session.execute(stmt)
            existing = result.scalars().first()

            if existing:
                existing.enabled = _parse_bool(row.get("启用"))
                existing.rate_type = _parse_str(row.get("评价类型")) or existing.rate_type
                existing.text_content = _parse_str(row.get("评价内容")) or existing.text_content
                existing.api_url = _parse_str(row.get("API地址")) or existing.api_url
            else:
                cfg = AutoRateConfig(
                    account_id=account_id,
                    enabled=_parse_bool(row.get("启用")),
                    rate_type=_parse_str(row.get("评价类型")) or "text",
                    text_content=_parse_str(row.get("评价内容")) or None,
                    api_url=_parse_str(row.get("API地址")) or None,
                )
                self.session.add(cfg)

        await self.session.commit()

    # ==================== 启动逻辑 ====================

    async def _get_accounts_to_start(self, enable_all: bool) -> list[tuple[str, str]]:
        """获取需要启动的账号列表 [(account_id, cookie)]"""
        stmt = select(XYAccount.account_id, XYAccount.cookie, XYAccount.status).where(
            XYAccount.owner_id == self.owner_id,
            XYAccount.account_id.in_(list(self._account_id_to_pk.keys())),
        )
        result = await self.session.execute(stmt)
        accounts_to_start = []
        for account_id, cookie, status in result.all():
            if not cookie or not cookie.strip():
                continue
            if enable_all or status == "active":
                accounts_to_start.append((account_id, cookie))
        return accounts_to_start

    async def _start_account(self, account_id: str, cookie: str) -> None:
        """启动单个账号的WebSocket任务"""
        from app.services.websocket_client import websocket_client
        result = await websocket_client.start_account(account_id, cookie, self.owner_id)
        if isinstance(result, dict) and not result.get("success", True):
            raise Exception(result.get("message", "启动失败"))
