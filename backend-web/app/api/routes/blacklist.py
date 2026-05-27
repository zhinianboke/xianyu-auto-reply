"""
黑名单管理路由

功能：
1. 个人黑名单CRUD（列表、新建、删除、启用/禁用、批量删除）
2. 个人黑名单导出/导入Excel
3. 闲鱼黑名单查询
"""
from __future__ import annotations

import io
import time

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from pydantic import BaseModel

from app.api import deps
from common.models.user import User
from common.utils.auth_scope import resolve_owner_scope
from app.services.blacklist_service import BlacklistService

router = APIRouter(prefix="/blacklist", tags=["黑名单管理"])


# ==================== 依赖注入 ====================

async def get_blacklist_service(session=Depends(deps.get_db_session)):
    return BlacklistService(session)


# ==================== 请求模型 ====================

class CreatePersonalBlacklistRequest(BaseModel):
    """新建个人黑名单请求"""
    account_id: str | None = None
    buyer_ids: str  # 英文逗号分隔
    item_id: str | None = None
    reason: str | None = None
    is_enabled: bool = True


class TogglePersonalBlacklistRequest(BaseModel):
    """启用/禁用请求"""
    is_enabled: bool


class BatchDeletePersonalBlacklistRequest(BaseModel):
    """批量删除请求"""
    ids: list[int]


# ==================== 个人黑名单接口 ====================
# 注意：固定路径路由必须在参数化路由（{record_id}）之前注册

@router.get("/personal")
async def list_personal_blacklist(
    buyer_id: str | None = Query(default=None, description="买家ID筛选"),
    buyer_nick: str | None = Query(default=None, description="买家昵称筛选"),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=100),
    current_user: User = Depends(deps.get_current_active_user),
    service: BlacklistService = Depends(get_blacklist_service),
):
    """获取个人黑名单列表"""
    owner_id, _ = resolve_owner_scope(current_user)

    items, total = await service.list_personal(
        owner_id=owner_id,
        buyer_id=buyer_id,
        buyer_nick=buyer_nick,
        page=page,
        page_size=page_size,
    )

    return {
        "success": True,
        "data": [
            {
                "id": item.id,
                "owner_id": item.owner_id,
                "account_id": item.account_id,
                "buyer_id": item.buyer_id,
                "buyer_nick": item.buyer_nick,
                "item_id": item.item_id,
                "reason": item.reason,
                "is_enabled": item.is_enabled,
                "created_at": item.created_at.isoformat() if item.created_at else None,
                "updated_at": item.updated_at.isoformat() if item.updated_at else None,
            }
            for item in items
        ],
        "total": total,
        "page": page,
        "page_size": page_size,
    }


@router.post("/personal")
async def create_personal_blacklist(
    request: CreatePersonalBlacklistRequest,
    current_user: User = Depends(deps.get_current_active_user),
    service: BlacklistService = Depends(get_blacklist_service),
):
    """新建个人黑名单（支持批量，自动判重）"""
    buyer_ids = [bid.strip() for bid in request.buyer_ids.split(",") if bid.strip()]
    if not buyer_ids:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="买家ID不能为空")

    created = await service.create_personal(
        owner_id=current_user.id,
        buyer_ids=buyer_ids,
        account_id=request.account_id,
        item_id=request.item_id,
        reason=request.reason,
        is_enabled=request.is_enabled,
    )

    skipped = len(buyer_ids) - len(created)
    msg = f"成功添加 {len(created)} 条黑名单"
    if skipped > 0:
        msg += f"，{skipped} 条已存在已跳过"

    return {
        "success": True,
        "message": msg,
        "data": {"count": len(created), "skipped": skipped},
    }


@router.post("/personal/batch-delete")
async def batch_delete_personal_blacklist(
    request: BatchDeletePersonalBlacklistRequest,
    current_user: User = Depends(deps.get_current_active_user),
    service: BlacklistService = Depends(get_blacklist_service),
):
    """批量删除个人黑名单"""
    if not request.ids:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="请选择要删除的记录")
    owner_id, _ = resolve_owner_scope(current_user)
    deleted = await service.batch_delete_personal(request.ids, owner_id)
    return {"success": True, "message": f"成功删除 {deleted} 条记录", "data": {"deleted": deleted}}


@router.get("/personal/export")
async def export_personal_blacklist(
    current_user: User = Depends(deps.get_current_active_user),
    service: BlacklistService = Depends(get_blacklist_service),
):
    """导出个人黑名单为Excel"""
    owner_id, _ = resolve_owner_scope(current_user)

    # 查询所有数据（不分页）
    items, _ = await service.list_personal(owner_id=owner_id, page=1, page_size=10000)

    workbook = Workbook()
    ws = workbook.active
    ws.title = "个人黑名单"

    headers = ["账号ID", "买家ID", "买家昵称", "商品ID", "拉黑原因", "是否启用"]
    ws.append(headers)

    for item in items:
        row_data = [
            item.account_id or "",
            item.buyer_id,
            item.buyer_nick or "",
            item.item_id or "",
            item.reason or "",
            "是" if item.is_enabled else "否",
        ]
        ws.append(row_data)

    # 将所有数据单元格设为文本格式，防止Excel把数字字符串转成数字
    for col_idx in range(1, len(headers) + 1):
        for row_idx in range(1, len(items) + 2):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.number_format = '@'
        # 自动列宽
        col_letter = get_column_letter(col_idx)
        max_len = len(headers[col_idx - 1])
        for row_idx in range(2, min(len(items) + 2, 50)):
            cell_val = ws.cell(row=row_idx, column=col_idx).value or ""
            max_len = max(max_len, min(len(str(cell_val)), 40))
        ws.column_dimensions[col_letter].width = max_len + 2

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f"personal_blacklist_{int(time.time())}.xlsx"
    return StreamingResponse(
        io.BytesIO(output.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post("/personal/import")
async def import_personal_blacklist(
    file: UploadFile = File(...),
    current_user: User = Depends(deps.get_current_active_user),
    service: BlacklistService = Depends(get_blacklist_service),
):
    """从Excel导入个人黑名单"""
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="请上传Excel文件(.xlsx或.xls)")

    contents = await file.read()
    try:
        workbook = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Excel文件读取失败: {str(exc)}") from exc

    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Excel文件为空")

    header = [str(cell).strip() if cell is not None else "" for cell in rows[0]]

    # 检查必要的列
    required_columns = ["买家ID"]
    missing_columns = [col for col in required_columns if col not in header]
    if missing_columns:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Excel缺少必要的列: {', '.join(missing_columns)}")

    # 列索引映射
    col_map = {name: header.index(name) for name in header if name}

    created_total = 0
    skipped_total = 0

    for row in rows[1:]:
        def get_cell(col_name: str, _row=row) -> str:
            idx = col_map.get(col_name)
            if idx is None or idx >= len(_row):
                return ""
            val = _row[idx]
            if val is None:
                return ""
            # 如果是浮点数（Excel把长数字字符串当数字读取），转为整数字符串
            if isinstance(val, float):
                if val == int(val):
                    return str(int(val))
                return str(val)
            s = str(val).strip()
            # 去除Excel数字后缀 .0
            if s.endswith(".0") and col_name in ("账号ID", "买家ID", "商品ID"):
                s = s[:-2]
            return s

        buyer_id = get_cell("买家ID")
        if not buyer_id:
            continue

        account_id = get_cell("账号ID") or None
        item_id = get_cell("商品ID") or None
        reason = get_cell("拉黑原因") or None
        is_enabled_str = get_cell("是否启用")
        is_enabled = is_enabled_str != "否"

        created = await service.create_personal(
            owner_id=current_user.id,
            buyer_ids=[buyer_id],
            account_id=account_id,
            item_id=item_id,
            reason=reason,
            is_enabled=is_enabled,
        )
        if created:
            created_total += len(created)
        else:
            skipped_total += 1

    msg = f"导入完成，新增 {created_total} 条"
    if skipped_total > 0:
        msg += f"，{skipped_total} 条已存在已跳过"

    return {"success": True, "message": msg, "data": {"created": created_total, "skipped": skipped_total}}


@router.delete("/personal/{record_id}")
async def delete_personal_blacklist(
    record_id: int,
    current_user: User = Depends(deps.get_current_active_user),
    service: BlacklistService = Depends(get_blacklist_service),
):
    """删除个人黑名单"""
    owner_id, _ = resolve_owner_scope(current_user)
    success = await service.delete_personal(record_id, owner_id)
    if not success:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="记录不存在或无权操作")
    return {"success": True, "message": "删除成功"}


@router.patch("/personal/{record_id}/toggle")
async def toggle_personal_blacklist(
    record_id: int,
    request: TogglePersonalBlacklistRequest,
    current_user: User = Depends(deps.get_current_active_user),
    service: BlacklistService = Depends(get_blacklist_service),
):
    """启用/禁用个人黑名单"""
    owner_id, _ = resolve_owner_scope(current_user)
    success = await service.toggle_personal(record_id, owner_id, request.is_enabled)
    if not success:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="记录不存在或无权操作")
    return {"success": True, "message": "操作成功"}


# ==================== 闲鱼黑名单接口 ====================

@router.get("/platform")
async def list_platform_blacklist(
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=100),
    current_user: User = Depends(deps.get_current_active_user),
    service: BlacklistService = Depends(get_blacklist_service),
):
    """获取闲鱼黑名单列表"""
    items, total = await service.list_platform(page=page, page_size=page_size)

    return {
        "success": True,
        "data": items,
        "total": total,
        "page": page,
        "page_size": page_size,
    }
