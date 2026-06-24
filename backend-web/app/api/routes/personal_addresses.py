"""
个人发布地址库接口

功能：
1. 提供个人地址库分页查询与增删改（每个用户仅能管理自己的数据）
2. 提供个人地址库 Excel 导入（按地址文本去重）与导出
3. 个人地址优先参与商品发布
"""
from __future__ import annotations

import io
import time
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, File, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import get_current_active_user, get_db_session
from common.models.user import User
from common.schemas.common import ApiResponse
from app.services.user_publish_address_service import UserPublishAddressService, _address_to_dict

router = APIRouter(prefix="/product-publish/personal-addresses", tags=["个人发布地址库"])


class PersonalAddressCreateRequest(BaseModel):
    """创建个人地址请求"""

    address: str = Field(..., min_length=1, max_length=200, description="地址文本")


class PersonalAddressUpdateRequest(BaseModel):
    """更新个人地址请求"""

    address: str = Field(..., min_length=1, max_length=200, description="地址文本")


class PersonalAddressBatchDeleteRequest(BaseModel):
    """批量删除个人地址请求"""

    ids: List[int] = Field(default_factory=list, description="个人地址ID列表")


@router.get("", response_model=ApiResponse)
async def list_personal_addresses(
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(20, description="每页条数"),
    keyword: Optional[str] = Query(None, description="关键词"),
    current_user: User = Depends(get_current_active_user),
    session: AsyncSession = Depends(get_db_session),
) -> Dict[str, Any]:
    """分页查询个人地址库"""
    svc = UserPublishAddressService(session)
    data = await svc.list_addresses(
        owner_id=current_user.id,
        page=page,
        page_size=page_size,
        keyword=keyword,
    )
    return ApiResponse(success=True, message="查询成功", data=data)


@router.post("", response_model=ApiResponse)
async def create_personal_address(
    req: PersonalAddressCreateRequest,
    current_user: User = Depends(get_current_active_user),
    session: AsyncSession = Depends(get_db_session),
) -> Dict[str, Any]:
    """创建个人地址"""
    svc = UserPublishAddressService(session)
    try:
        address = await svc.create(current_user.id, req.address)
    except ValueError as exc:
        return ApiResponse(success=False, message=str(exc))
    return ApiResponse(success=True, message="个人地址创建成功", data={"address": _address_to_dict(address)})


@router.put("/{address_id}", response_model=ApiResponse)
async def update_personal_address(
    address_id: int,
    req: PersonalAddressUpdateRequest,
    current_user: User = Depends(get_current_active_user),
    session: AsyncSession = Depends(get_db_session),
) -> Dict[str, Any]:
    """更新个人地址"""
    svc = UserPublishAddressService(session)
    try:
        updated = await svc.update(current_user.id, address_id, req.address)
    except ValueError as exc:
        return ApiResponse(success=False, message=str(exc))
    if not updated:
        return ApiResponse(success=False, message="个人地址不存在")
    return ApiResponse(success=True, message="个人地址更新成功", data={"address": _address_to_dict(updated)})


@router.post("/batch-delete", response_model=ApiResponse)
async def batch_delete_personal_addresses(
    req: PersonalAddressBatchDeleteRequest,
    current_user: User = Depends(get_current_active_user),
    session: AsyncSession = Depends(get_db_session),
) -> Dict[str, Any]:
    """批量删除个人地址"""
    svc = UserPublishAddressService(session)
    try:
        success_count = await svc.batch_delete(current_user.id, req.ids)
    except ValueError as exc:
        return ApiResponse(success=False, message=str(exc))

    return ApiResponse(
        success=True,
        message=f"成功删除 {success_count} 条个人地址",
        data={"success_count": success_count, "total_count": len(req.ids)},
    )


@router.get("/export")
async def export_personal_addresses(
    current_user: User = Depends(get_current_active_user),
    session: AsyncSession = Depends(get_db_session),
):
    """导出个人地址库为Excel"""
    svc = UserPublishAddressService(session)
    records = await svc.list_all_for_owner(current_user.id)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "个人地址库"

    headers = ["地址"]
    worksheet.append(headers)
    for record in records:
        worksheet.append([record.address or ""])

    # 所有单元格设为文本格式，并自适应列宽
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(headers[col_idx - 1])
        for row_idx in range(1, len(records) + 2):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.number_format = '@'
            if row_idx >= 2:
                cell_val = cell.value or ""
                max_len = max(max_len, min(len(str(cell_val)), 60))
        worksheet.column_dimensions[col_letter].width = max_len + 2

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f"personal_addresses_{int(time.time())}.xlsx"
    return StreamingResponse(
        io.BytesIO(output.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post("/import", response_model=ApiResponse)
async def import_personal_addresses(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_active_user),
    session: AsyncSession = Depends(get_db_session),
) -> Dict[str, Any]:
    """从Excel导入个人地址库（按地址文本去重，更新或插入）"""
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        return ApiResponse(success=False, message="请上传Excel文件(.xlsx或.xls)")

    contents = await file.read()
    try:
        workbook = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    except Exception as exc:
        return ApiResponse(success=False, message=f"Excel文件读取失败: {str(exc)}")

    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return ApiResponse(success=False, message="Excel文件为空")

    header = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    if "地址" not in header:
        return ApiResponse(success=False, message="Excel缺少必要的列: 地址")

    address_idx = header.index("地址")
    addresses: List[str] = []
    for row in rows[1:]:
        if address_idx >= len(row):
            continue
        value = row[address_idx]
        if value is None:
            continue
        text = str(value).strip()
        if text:
            addresses.append(text)

    if not addresses:
        return ApiResponse(success=False, message="未读取到有效的地址数据")

    svc = UserPublishAddressService(session)
    result = await svc.upsert_many(current_user.id, addresses)
    created = result.get("created", 0)
    updated = result.get("updated", 0)
    return ApiResponse(
        success=True,
        message=f"导入完成，新增 {created} 条，更新 {updated} 条",
        data=result,
    )
