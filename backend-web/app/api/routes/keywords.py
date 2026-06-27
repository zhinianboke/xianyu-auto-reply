from __future__ import annotations

import io
import time

from openpyxl import Workbook, load_workbook
from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse

from app.api import deps
from common.models.user import User
from common.schemas.common import ApiResponse
from common.schemas.keyword import KeywordDetail, KeywordTextList, KeywordTextUpdatePayload
from common.utils.auth_scope import resolve_owner_scope
from common.utils.local_image_upload import ImageUploadError, save_uploaded_image
from app.services.account_service import AccountService
from app.services.keyword_service import KeywordService

router = APIRouter(tags=["keywords"])


@router.get("", response_model=list[KeywordDetail])
async def get_all_keywords(
    current_user: User = Depends(deps.get_current_active_user),
    keyword_service: KeywordService = Depends(deps.get_keyword_service),
) -> list[KeywordDetail]:
    """获取当前用户所有账号的关键词列表，管理员可查看所有"""
    owner_id, _ = resolve_owner_scope(current_user)
    keywords = await keyword_service.list_keywords_for_owner(owner_id)
    return [KeywordDetail(**item) for item in keywords]


@router.get("/{account_id}", response_model=list[KeywordDetail])
async def get_keywords_with_item_id(
    account_id: str,
    current_user: User = Depends(deps.get_current_active_user),
    account_service: AccountService = Depends(deps.get_account_service),
    keyword_service: KeywordService = Depends(deps.get_keyword_service),
) -> list[KeywordDetail]:
    """获取关键词列表，管理员可查看所有账号"""
    _, is_admin = resolve_owner_scope(current_user)

    # 管理员可以查看任意账号，普通用户只能查看自己的
    if is_admin:
        account = await account_service.get_account_by_identifier(account_id)
    else:
        account = await account_service.get_account_for_user(current_user.id, account_id)
    
    if not account:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="账号不存在")
    keywords = await keyword_service.list_keywords(account)
    return [KeywordDetail(**item) for item in keywords]


@router.post("/{account_id}", response_model=ApiResponse)
async def save_keywords_with_item_id(
    account_id: str,
    payload: KeywordTextList,
    current_user: User = Depends(deps.get_current_active_user),
    account_service: AccountService = Depends(deps.get_account_service),
    keyword_service: KeywordService = Depends(deps.get_keyword_service),
) -> ApiResponse:
    # 管理员可以操作所有账号，普通用户只能操作自己的账号
    owner_id, _ = resolve_owner_scope(current_user)
    account = await account_service.get_account_for_user(owner_id, account_id)
    if not account:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="账号不存在")
    try:
        await keyword_service.replace_text_keywords(
            account,
            [entry.model_dump() for entry in payload.keywords],
        )
    except ValueError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc
    return ApiResponse(success=True, message="关键词保存成功")


@router.put("/{account_id}/{keyword}", response_model=ApiResponse)
async def update_single_keyword(
    account_id: str,
    keyword: str,
    payload: KeywordTextUpdatePayload,
    old_item_id: str = "",
    current_user: User = Depends(deps.get_current_active_user),
    account_service: AccountService = Depends(deps.get_account_service),
    keyword_service: KeywordService = Depends(deps.get_keyword_service),
) -> ApiResponse:
    owner_id, _ = resolve_owner_scope(current_user)
    source_account = await account_service.get_account_for_user(owner_id, account_id)
    if not source_account:
        return ApiResponse(success=False, message="原所属账号不存在")

    target_account = await account_service.get_account_for_user(owner_id, payload.account_id)
    if not target_account:
        return ApiResponse(success=False, message="目标所属账号不存在")

    try:
        await keyword_service.update_text_keyword(
            source_account=source_account,
            target_account=target_account,
            source_keyword=keyword,
            source_item_id=old_item_id or None,
            target_keyword=payload.keyword,
            target_reply=payload.reply,
            target_item_id=payload.item_id,
        )
    except ValueError as exc:
        return ApiResponse(success=False, message=str(exc))

    return ApiResponse(success=True, message="关键词更新成功")


# ==================== 关键词导入导出 ====================

@router.get("/{account_id}/export")
async def export_keywords(
    account_id: str,
    current_user: User = Depends(deps.get_current_active_user),
    account_service: AccountService = Depends(deps.get_account_service),
    keyword_service: KeywordService = Depends(deps.get_keyword_service),
):
    """导出指定账号的关键词为Excel文件"""
    # 管理员可以操作所有账号，普通用户只能操作自己的账号
    owner_id, _ = resolve_owner_scope(current_user)
    account = await account_service.get_account_for_user(owner_id, account_id)
    if not account:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="账号不存在")
    
    keywords = await keyword_service.list_keywords(account)
    
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "关键词数据"
    worksheet.append(["关键词", "商品ID", "关键词内容"])

    for kw in keywords:
        if kw.get("type", "text") == "text":
            worksheet.append([
                kw["keyword"],
                kw.get("item_id") or "",
                kw["reply"],
            ])

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    
    filename = f"keywords_{account_id}_{int(time.time())}.xlsx"
    
    return StreamingResponse(
        io.BytesIO(output.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.post("/{account_id}/import")
async def import_keywords(
    account_id: str,
    file: UploadFile = File(...),
    current_user: User = Depends(deps.get_current_active_user),
    account_service: AccountService = Depends(deps.get_account_service),
    keyword_service: KeywordService = Depends(deps.get_keyword_service),
):
    """导入Excel文件中的关键词到指定账号"""
    # 管理员可以操作所有账号，普通用户只能操作自己的账号
    owner_id, _ = resolve_owner_scope(current_user)
    account = await account_service.get_account_for_user(owner_id, account_id)
    if not account:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="账号不存在")
    
    # 检查文件类型
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="请上传Excel文件(.xlsx或.xls)")
    
    # 读取Excel文件
    contents = await file.read()
    try:
        workbook = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Excel文件读取失败: {str(exc)}") from exc

    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Excel文件为空")

    header = [str(cell).strip() if cell is not None else "" for cell in rows[0]]

    # 检查必要的列
    required_columns = ["关键词", "商品ID", "关键词内容"]
    missing_columns = [col for col in required_columns if col not in header]
    if missing_columns:
        raise HTTPException(status_code=400, detail=f"Excel文件缺少必要的列: {', '.join(missing_columns)}")

    column_index = {name: header.index(name) for name in required_columns}

    # 处理导入数据
    import_data = []
    for row in rows[1:]:
        keyword_cell = row[column_index["关键词"]] if len(row) > column_index["关键词"] else None
        item_id_cell = row[column_index["商品ID"]] if len(row) > column_index["商品ID"] else None
        reply_cell = row[column_index["关键词内容"]] if len(row) > column_index["关键词内容"] else None

        keyword = str(keyword_cell).strip() if keyword_cell is not None else ""
        item_id = str(item_id_cell).strip() if item_id_cell is not None else ""
        reply = str(reply_cell).strip() if reply_cell is not None else ""

        if item_id.endswith(".0"):
            item_id = item_id[:-2]

        if not keyword:
            continue

        import_data.append({
            "keyword": keyword,
            "reply": reply,
            "item_id": item_id
        })
    
    if not import_data:
        raise HTTPException(status_code=400, detail="Excel文件中没有有效的关键词数据")
    
    # 保存到数据库
    await keyword_service.replace_text_keywords(account, import_data)
    
    return ApiResponse(
        success=True,
        message="导入成功",
        data={
            "added": len(import_data),
            "updated": 0
        }
    )


@router.post("/{account_id}/image")
async def add_image_keyword(
    account_id: str,
    keyword: str = Form(...),
    item_id: str = Form(default=""),
    image: UploadFile = File(...),
    current_user: User = Depends(deps.get_current_active_user),
    account_service: AccountService = Depends(deps.get_account_service),
    keyword_service: KeywordService = Depends(deps.get_keyword_service),
):
    """添加图片关键词"""
    # 管理员可以操作所有账号，普通用户只能操作自己的账号
    owner_id, _ = resolve_owner_scope(current_user)
    account = await account_service.get_account_for_user(owner_id, account_id)
    if not account:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="账号不存在")
    
    if not keyword or not keyword.strip():
        raise HTTPException(status_code=400, detail="关键词不能为空")
    
    if not image or not image.filename:
        raise HTTPException(status_code=400, detail="请选择图片文件")

    # 使用统一的上传目录
    from app.core.paths import get_upload_path
    upload_dir = get_upload_path("keywords")

    try:
        # 保留原行为：只校验类型，不限制文件大小
        filepath, filename, _ = await save_uploaded_image(
            image,
            upload_dir,
            validate_size=False,
        )
    except ImageUploadError as exc:
        raise HTTPException(status_code=exc.status_code, detail=exc.message)

    image_url = f"/static/uploads/keywords/{filename}"

    # 保存图片关键词到数据库
    # 空字符串统一转为 None，与 service 层保持一致
    normalized_item_id = item_id.strip() if item_id and item_id.strip() else None

    try:
        await keyword_service.add_image_keyword(
            account,
            keyword.strip(),
            image_url,
            normalized_item_id,
        )
    except ValueError as e:
        # 关键词已存在等业务错误：回滚已上传文件
        if filepath.exists():
            filepath.unlink()
        return ApiResponse(success=False, message=str(e))
    except Exception as e:
        # 其他异常：回滚已上传文件
        if filepath.exists():
            filepath.unlink()
        return ApiResponse(success=False, message=f"图片关键词保存失败: {str(e)}")

    return ApiResponse(
        success=True,
        message="图片关键词添加成功",
        data={
            "keyword": keyword,
            "image_url": image_url,
            "item_id": normalized_item_id,
        },
    )


@router.delete("/{account_id}/{keyword}")
async def delete_single_keyword(
    account_id: str,
    keyword: str,
    item_id: str = "",
    rule_id: int | None = None,
    current_user: User = Depends(deps.get_current_active_user),
    account_service: AccountService = Depends(deps.get_account_service),
    keyword_service: KeywordService = Depends(deps.get_keyword_service),
) -> ApiResponse:
    """删除单个关键词（支持文本和图片类型）"""
    # 管理员可以操作所有账号，普通用户只能操作自己的账号
    owner_id, _ = resolve_owner_scope(current_user)
    account = await account_service.get_account_for_user(owner_id, account_id)
    if not account:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="账号不存在")
    
    try:
        deleted = await keyword_service.delete_keyword(account, keyword, item_id or None, rule_id=rule_id)
        if not deleted:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="关键词不存在")
        return ApiResponse(success=True, message="关键词删除成功")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"删除失败: {str(e)}")
